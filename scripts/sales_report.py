'''
Sales Report Automation is a Python application designed to process, aggregate, 
and report employee data from various Excel files, manage filesystem directories, 
and interact with Microsoft Outlook for report distribution. 

## Features

- Reading and processing Excel files with employee data.
- Aggregating employee sales and work hours data.
- Configuration management using INI files (sales.ini).
- Generating a master report and updating Excel workbooks.
- Interacting with Microsoft Outlook to send emails with attachments.
- Logging errors and ensuring the necessary filesystem structure.
- Sample data can be found the sample_data\sales folder.

## Prerequisites

Before you begin, ensure you have met the following requirements:
- Python 3.x installed
- Microsoft Outlook installed (for Outlook features)
- Python libraries: `pandas`, `openpyxl`, `win32com`, `configparser`

'''

# Imports
import pandas as pd
import win32com.client as win32
import configparser
import os
import shutil
from datetime import date
from typing import List, Dict, Optional
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Helper functions
def ensure_folder_exists(folder_name: str) -> None: 
    """Ensure that a folder exists; if not, create it."""
    try:
        os.makedirs(folder_name, exist_ok=True)
    except OSError as e:
        print(f"An error occurred while creating the folder: {e}")

def exit_program(message: str) -> None:
    """Exit the program after waiting for user input."""
    print(message)
    input("Press any key to exit...")
    raise SystemExit

def identify_week_num(input_date: date) -> int:
    """Returns the week number within the current month. Week 1 starts on the 1st of the month."""
    first_day_of_month = input_date.replace(day=1)
    first_week_day = first_day_of_month.weekday()
    return ((input_date.day + first_week_day - 1) // 7) + 1

def identify_month_name(input_date: date) -> str:
    """Identify the month name for a given date."""
    return input_date.strftime('%B')
    
def read_excel_file(folder_directory: str, file_name: str, sheetname: Optional[str] = None) -> pd.DataFrame:
    """Read a pending report file and return its data."""
    if sheetname is None:
        return pd.read_excel(f'{folder_directory}/{file_name}')
    return pd.read_excel(f'{folder_directory}/{file_name}', sheet_name=sheetname)

def log_error(error_message: str) -> None:
    """Log an error message to a file."""
    #check if file exists
    if not os.path.exists("error_log.txt"):
        with open("error_log.txt", "w") as log_file:
            log_file.write(f"An error occurred: {error_message}\n")
    else:
        with open("error_log.txt", "a") as log_file:
            log_file.write(f"An error occurred: {error_message}\n")

# Classes

# Configuration Handling
class Config:
    def __init__(self, filepath):
        """
        Initialize the configuration handler.

        Reads the configuration from the specified ini file and provides
        methods to access various configuration values such as folder paths
        and file names.

        Args:
        filepath (str): The path to the configuration file.
        """
        self.config = configparser.ConfigParser()
        self.config.read(filepath)
    
    def get_folder_path(self, folder_name: str) -> str:
        return self.config['FOLDERS'][folder_name]
    
    def get_file_name(self, file_name: str) -> str:
        return self.config['FILES'][file_name]

    def get_outlook_config(self) -> Dict[str, str]:
        outlook_config = {k.upper(): v for k, v in self.config['OUTLOOK'].items()}
        outlook_config['OUTLOOK_EMAIL_ATTACHMENT'] = f"{self.get_folder_path('MASTER_FOLDER_PATH')}/{self.get_file_name('MASTER_FILE_NAME')}"
        return outlook_config

# Filesystem Handling
class FilesystemPreparer:
    def __init__(self, config):
        """
        Prepare the filesystem based on the provided configuration.

        This class ensures that all necessary folders are present and accessible
        before the program proceeds with its operations.

        Args:
        config (Config): The configuration handler object with methods to retrieve
        configuration values.
        """
        self.config = config
        self.configure_folders()

    def configure_folders(self) -> None:
        """Configure necessary folders for the program."""
        ensure_folder_exists(self.config.get_folder_path('PENDING_FOLDER_NAME'))
        ensure_folder_exists(self.config.get_folder_path('COMPLETED_FOLDER_NAME'))
        ensure_folder_exists(self.config.get_folder_path('MASTER_FOLDER_PATH'))

# Report Handling
class PendingReports():
    def __init__(self, config):
        """
        Process pending report files and prepare them for inclusion in the master report.

        Inherits from FilesystemPreparer to ensure all directories are properly set up,
        then proceeds to identify and process all pending reports, aggregating necessary
        employee data.

        Args:
        config (Config): The configuration handler object which provides access
        to file paths and other settings.
        """
        self.config = config
        self.month_name = set()
        self.pending_files = self.check_files()
        self.employee_data_list = self.process_pending_reports()
        self.pending_data = self.combine_dataframes()
        
    def check_files(self) -> List[str]:
        """Check if the necessary files exist and return a list of pending files."""
        master_file_path = f"{self.config.get_folder_path('MASTER_FOLDER_PATH')}/{self.config.get_file_name('MASTER_FILE_NAME')}"
        pending_folder_name = self.config.get_folder_path('PENDING_FOLDER_NAME')
        if not os.path.exists(master_file_path):
            exit_program("The master spreadsheet is missing. Exiting...")
        if not os.listdir(pending_folder_name):
            exit_program("There are no pending reports to add to the master spreadsheet. Exiting...")
        return os.listdir(pending_folder_name)

    def process_pending_reports(self) -> List[pd.DataFrame]:
        """Process all pending reports and return a list of employee data."""
        employee_data_list = []
        for file_name in self.pending_files:
            employee_data = read_excel_file(self.config.get_folder_path('PENDING_FOLDER_NAME'), file_name)
            aggregated_data = self.aggregate_data(employee_data, file_name)
            employee_data_list.append(aggregated_data)
        self.move_completed_files()
        return employee_data_list

    def aggregate_data(self, employee_data: pd.DataFrame, file_name: str) -> pd.DataFrame:
        """Aggregate data from employee reports and return a dataframe."""
        newest_date = self.collect_month(employee_data)
        employee_dict = {}
        employee_dict['Week'] = identify_week_num(newest_date)
        employee_dict['Employee ID'] = self.get_employee_id(file_name)
        employee_dict['Employee Name'] = self.get_employee_name(employee_data)
        employee_dict['Hours Worked'] = self.get_hours_worked(employee_data)
        employee_dict['Sales'] = self.get_total_sales(employee_data)
        employee_dict = pd.DataFrame(employee_dict, index=[0])
        return employee_dict
    
    def combine_dataframes(self) -> pd.DataFrame:
        """Combine all employee dataframes into a single dataframe."""
        return pd.concat(self.employee_data_list)

    def collect_month(self, employee_data: pd.DataFrame) -> None:
        """Collect the month name from the employee data."""
        newest_date = self.get_newest_date(employee_data)
        month_data = identify_month_name(newest_date)
        if len(self.month_name) < 1:
            self.month_name.add(month_data)
            self.month_number = newest_date.month
            self.year_number = newest_date.year
            self.week_number = identify_week_num(newest_date)
        if month_data in self.month_name:
            return newest_date
        raise ValueError(f"Month '{month_data}' is not the same as the other months in the report. Please fix this and try again. Exiting...")
    
    def move_completed_files(self) -> None:
        """Move completed files to the completed folder."""
        for file_name in self.pending_files:
            shutil.move(f"{self.config.get_folder_path('PENDING_FOLDER_NAME')}/{file_name}", f"{self.config.get_folder_path('COMPLETED_FOLDER_NAME')}/{self.year_number}_{self.month_number}_week{self.week_number}_{file_name}")
            print(f"File '{file_name}' moved to the completed folder.")

    @staticmethod
    def get_employee_id(file_name) -> int:
        """Get the employee ID from the file name."""
        try:
            employee_id = file_name.split(' ')[-1].split('.')[0]
            return int(employee_id)
        except ValueError:
            raise ValueError(f"Employee ID in '{file_name}' is not a valid number.")

    @staticmethod
    def get_hours_worked(employee_data: pd.DataFrame) -> int:
        """Get the total hours worked from the employee data."""
        return employee_data['Hours Worked'].sum()

    @staticmethod
    def get_newest_date(employee_data: pd.DataFrame) -> date:
        """Get the newest date from the employee data."""
        return employee_data['Date'].max()

    @staticmethod
    def get_employee_name(employee_data: pd.DataFrame) -> str:
        """Get the employee name from the employee data, ensure there's only one unique name."""
        if len(employee_data['Employee Name'].unique()) > 1:
            exit_program("There are multiple employee names in the report. Please fix this and try again. Exiting...")
        return employee_data['Employee Name'].iloc[0]
    
    @staticmethod
    def get_total_sales(employee_data: pd.DataFrame) -> int:
        """Get the total sales from the employee data."""
        return employee_data['Sales'].sum()

# Master Report Handling
class MasterReport(PendingReports):
    def __init__(self,config):
        """
        Handle the creation and updating of the master report.

        This class is responsible for adding new data from processed pending reports
        to the master report file, ensuring no duplicates and that the data is saved
        correctly.

        Args:
        config (Config): The configuration handler object used to manage file paths
        and settings throughout the class.
        """
        super().__init__(config)
        self.current_month = list(self.month_name)[0]
        self.master_file = read_excel_file(self.config.get_folder_path('MASTER_FOLDER_PATH'), self.config.get_file_name('MASTER_FILE_NAME'), self.current_month)
        self.add_employee_data()
    
    def add_employee_data(self) -> None:
        """Adds only new employee data to the master file."""
        self.new_data = self.exclude_existing_records()
        self.master_file = pd.concat([self.master_file, self.new_data])
        self.save_to_excel()
    
    def exclude_existing_records(self) -> pd.DataFrame:
        """Check for unique rows in employee_data_list that are not in master_file and return them."""
        merged = self.master_file.merge(self.pending_data, 
                                        on=self.pending_data.columns.tolist(), 
                                        how='outer', 
                                        indicator=True)
        unique_to_employee_data = merged[merged['_merge'] == 'right_only']
        unique_to_employee_data = unique_to_employee_data.drop(columns=['_merge'])
        return unique_to_employee_data
    
    def save_to_excel(self) -> None:
        """Save the master report to an Excel file."""
        file_path = f"{self.config.get_folder_path('MASTER_FOLDER_PATH')}/{self.config.get_file_name('MASTER_FILE_NAME')}"
        try:
            workbook = load_workbook(file_path)
            sheet_names = workbook.sheetnames
            dfs = {}
            for sheet_name in sheet_names:
                dfs[sheet_name] = pd.read_excel(file_path, sheet_name=sheet_name)
            dfs[self.current_month] = self.master_file
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, df in dfs.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            self.format_excel_file(file_path)
            print(f"Master file saved to {file_path}")
        except Exception as e:
            print(f"An error occurred while saving the file: {e}")
    
    @staticmethod
    def format_excel_file(file_path: str) -> None:
        workbook = load_workbook(file_path)
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            worksheet.auto_filter.ref = worksheet.dimensions
            
            for column_cells in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in column_cells if cell.value) + 8
                worksheet.column_dimensions[column_cells[0].column_letter].width = max_length
                for cell in column_cells:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        workbook.save(file_path)
        
# Outlook Handling
class Outlook:
    def __init__(self, config):
        """
        Handle Outlook interactions for sending and receiving reports.

        This class sets up the necessary configuration for interacting with Outlook,
        allows for downloading attachments from specified email folders, and sends
        out emails with the master report as an attachment.

        Args:
        config (Config): The configuration handler object which provides access
        to Outlook settings and other configurations.
        """
        self.config = config
        self.config_outlook = self.config.get_outlook_config()
        self.email_to = self.config_outlook['OUTLOOK_EMAIL_TO']
        self.email_cc = self.config_outlook['OUTLOOK_EMAIL_CC']
        self.email_bcc = self.config_outlook['OUTLOOK_EMAIL_BCC']
        self.email_subject = self.config_outlook['OUTLOOK_EMAIL_SUBJECT']
        self.email_body = self.config_outlook['OUTLOOK_EMAIL_BODY']
        self.attachment_path = self.config_outlook['OUTLOOK_EMAIL_ATTACHMENT']
        self.folder_name = self.config_outlook['OUTLOOK_FOLDER_NAME']
        self.archive_folder_name = self.config_outlook['OUTLOOK_ARCHIVE_FOLDER_NAME']
        self.pending_folder_name = self.config.get_folder_path('PENDING_FOLDER_NAME')
        
    def send_master_reports(self) -> None:
        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = self.email_to
        mail.CC = self.email_cc
        mail.BCC = self.email_bcc
        mail.Subject = self.email_subject
        mail.Body = self.email_body
        mail.Attachments.Add(os.path.abspath(self.attachment_path))
        mail.Display()
        print(f"Draft email created with '{self.attachment_path}' attached!")

    def download_outlook_attachments(self) -> None:
        outlook = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")
        inbox = outlook.GetDefaultFolder(6)
        try:
            archive_folder = inbox.Folders[self.archive_folder_name]
        except KeyError:
            print(f"The folder '{self.archive_folder_name}' does not exist in the Inbox.")
        try:
            weekly_reports_folder = inbox.Folders[self.folder_name]
        except KeyError:
            print(f"The folder '{self.folder_name}' does not exist in the Inbox.")
        if weekly_reports_folder.Items.Count == 0:
            print(f"The folder '{self.folder_name}' is empty.")
        else:
            for message in weekly_reports_folder.Items:
                for attachment in message.Attachments:
                    attachment.SaveAsFile(os.path.abspath(f'{self.pending_folder_name}/{attachment.FileName}'))
                    print(f"Attachment '{attachment.FileName}' downloaded from Outlook.")
                message.Move(archive_folder)
# Main
if __name__ == "__main__":
    try:
        config_file = 'config/sales.ini'
        config = Config(config_file)
        outlook = Outlook(config)
        outlook.download_outlook_attachments()
        master_report = MasterReport(config)
        outlook.send_master_reports()
    except Exception as e:
        log_error(e)
        print(f"An unexpected error occurred: {e}. Please check the error_log.txt for more details.")